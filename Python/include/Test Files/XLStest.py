# Write to and display a text file
from openpyxl import load_workbook, Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from tkinter import *

from guizero import App, Text, Box
import time
import easygui
from multiprocessing import Queue
# import outputfile


targetTimes = [3000, 4500, 2960, 6200, 5754, 7430]
ballSpeeds = [4.44, 2.52, 4.32, 3.2, 5.4, 10.2]
filename = "demo_wb.xlsx"
player = "Player2"
difficulty = 1
drillSpeed = 1
drillType = "Dynamic"
ball=[0,0,0,0,0,0]
data = Queue()

# result = targetTimes + ballSpeeds
# print(result)
# data.put(result)
#
# out = data.get()
# print(out)
# outputfile.OutPut(targetTimes,ballSpeeds, player).run()
#
# while True:
#     print("Hello")
try:
    wb = load_workbook(filename=filename)
    print("[Launcher]: file exists")
    new_book = False
except Exception as e:
    print("[Launcher]: file not found, making new book ")
    wb = Workbook()
    summary_sheet = wb.get_sheet_by_name('Sheet')
    summary_sheet.title = 'Summary'
    new_book = True

try:
    players_sheet = wb[player]
    print("[Launcher]: sheet exists, appending data")

except KeyError as k:
    print("[Launcher]: creating new sheet for: ", player)
    players_sheet = wb.create_sheet(title=player)

data_printed = False
col = 1
row = 1
drill = 1
while not data_printed:
    if players_sheet.cell(column=col,row=row+1).value is None:
        players_sheet.cell(column=col, row=row+1).value = "Drill"+str(drill)
        if row == 1:
            # players_sheet.cell(column=col, row=row).value = "."
            players_sheet.cell(column=col+1, row=row).value = "Ball #"
            players_sheet.cell(column=col+2, row=row).value = "Time for Pass"
            players_sheet.cell(column=col+3, row=row).value = "Ball Speed"
        for i in range(5):
            players_sheet.cell(column=col+1, row=row+i+1).value = row+i
            players_sheet.cell(column=col+2, row=row+i+1).value = (targetTimes[i+1]/1000)
            players_sheet.cell(column=col+3, row=row+i+1).value = ballSpeeds[i+1]

        data_printed = True

    else:
        row+=5
        drill +=1

wb.save(filename)

# PLOT THE DATA:
ranges = [0]*(len(wb.worksheets)-1)
speed_ranges = [0]*(len(wb.worksheets)-1)
series = [0]*20
speed_series = [0]*(len(wb.worksheets)-1)

chart1 = ScatterChart()
chart2 = ScatterChart()
chart3 = ScatterChart()
chart4 = ScatterChart()
chart5 = ScatterChart()
speed_chart1 = ScatterChart()
speed_chart2 = ScatterChart()
speed_chart3 = ScatterChart()
speed_chart4 = ScatterChart()
speed_chart5 = ScatterChart()

charts = [chart1,chart2,chart3,chart4,chart5]
speed_charts = [speed_chart1,speed_chart2,speed_chart3,speed_chart4,speed_chart5]
# otherchart = ScatterChart()
# speed_charts = [otherchart] * (len(wb.worksheets) - 1)

for i in range((len(wb.worksheets)-1)):
    print(str(wb.worksheets[i+1])[11:20])
    gathered = False
    coll=1
    roww=1
    drill=1
    charts[i].title = str(wb.worksheets[i+1])[12:19]+"'s  Results"
    speed_charts[i].title = str(wb.worksheets[i+1])[12:19]+"'s  Results"

    while not gathered:
        print("Gathering..")
        if wb.worksheets[i+1].cell(column=coll,row=roww+1).value is None:
            gathered = True
        else:
            roww+=5
            gathered = False

    ranges[i] = Reference(wb.worksheets[i+1],min_col=3,min_row=1,max_col=3,max_row=roww)
    x_vals = Reference(wb.worksheets[i+1],min_col=2,min_row=1,max_row=roww)
    series[i]=Series(values=ranges[i],xvalues=x_vals,title_from_data=True)
    charts[i].append(series[0])

    charts[i].x_axis.title = "Test Number"
    charts[i].x_axis.scaling.min = 0
    charts[i].x_axis.scaling.max = roww
    charts[i].y_axis.title = "Reaction Times (seconds)"
    charts[i].legend = None

    s1 = charts[i].series[0]
    s1.marker.symbol = "triangle"
    s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline

    s1.graphicalProperties.line.noFill = True

    speed_ranges[i] = Reference(wb.worksheets[i + 1], min_col=4, min_row=1, max_col=4, max_row=roww)
    x_value = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)
    series[i+10] = Series(values=speed_ranges[i], xvalues=x_value, title_from_data=True)
    speed_charts[i].append(series[i+10])

    speed_charts[i].x_axis.title = "Test Number"
    speed_charts[i].x_axis.scaling.min = 0
    speed_charts[i].x_axis.scaling.max = roww
    speed_charts[i].y_axis.title = "Ball Speeds (m/s)"
    speed_charts[i].legend = None

    s2 = speed_charts[i].series[0]
    s2.marker.symbol = "circle"
    s2.marker.graphicalProperties.solidFill = "006868"  # Marker filling
    s2.marker.graphicalProperties.line.solidFill = "006868"  # Marker outline

    s2.graphicalProperties.line.noFill = True

    charts[i].style = 13
    wb.worksheets[0].add_chart(charts[i], "A"+str(i+1 +15*i))
    speed_charts[i].style = 13
    wb.worksheets[0].add_chart(speed_charts[i], "J" + str(i + 1 + 15 * i))
# Style the lines
wb.save(filename)

# speed_ranges = [0]*(len(wb.worksheets)-1)
# listSpeeds = [0]*(len(wb.worksheets)-1)
# otherchart = ScatterChart()
# speed_charts = [otherchart] * (len(wb.worksheets) - 1)

# for i in range((len(wb.worksheets)-1)):
#     gathered = False
#     coll = 1
#     roww = 1
#     drill = 1
#
#     speed_charts[i].title = str(player) + "'s  Results"
#
#     while not gathered:
#         print("Gathering..")
#         if wb.worksheets[i + 1].cell(column=coll, row=roww + 1).value is None:
#             gathered = True
#         else:
#             roww += 5
#             gathered = False
#
#     speed_ranges[i] = Reference(wb.worksheets[i + 1], min_col=4, min_row=1,max_row=roww)
#     print(speed_ranges[i])
#     x_vals = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)
#
#     listSpeeds[i] = Series(values=speed_ranges[i], xvalues=x_vals, title_from_data=True)
#
#     # print(listSpeeds[i])
#     speed_charts[i].append(listSpeeds[0])
#     speed_charts[i].x_axis.title = "Test Number"
#     speed_charts[i].x_axis.scaling.min = 0
#     speed_charts[i].x_axis.scaling.max = roww
#     speed_charts[i].y_axis.title = "Ball Speeds (m/s)"
#     speed_charts[i].legend = None
#
#     s2 = speed_charts[i].listSpeeds[0]
#     s2.marker.symbol = "circle"
#     s2.marker.graphicalProperties.solidFill = "006868"  # Marker filling
#     s2.marker.graphicalProperties.line.solidFill = "006868"  # Marker outline
#
#     s2.graphicalProperties.line.noFill = True
#
#     speed_charts[i].style = 13
#     wb.worksheets[0].add_chart(speed_charts[i], "J" + str(i + 1 + 15 * i))
#
# wb.save(filename)

# for i in range(5):
#     ball[i+1] = "Pass # " + str(i + 1) + ":   " + str((targetTimes[i + 1]) / 1000) + " sec  |  " + str(ballSpeeds[i + 1]) + " m/s"
#
# results = Tk()
# var1 = StringVar()
# var2 = StringVar()
# var3 = StringVar()
# var4 = StringVar()
# var5 = StringVar()
#
#
# ball_1 = Message( results, width=500, textvariable = var1 )
# ball_2 = Message( results, width=500, textvariable = var2)
# ball_3 = Message( results, width=500, textvariable = var3)
# ball_4 = Message( results, width=500, textvariable = var4)
# ball_5 = Message( results, width=500, textvariable = var5)
#
#
# var1.set(ball[1])
# var2.set(ball[2])
# var3.set(ball[3])
# var4.set(ball[4])
# var5.set(ball[5])
#
# ball_1.pack()
# ball_2.pack()
# ball_3.pack()
# ball_4.pack()
# ball_5.pack()
#
#
# results.mainloop()
#
#
# easygui.msgbox(ball[1]+
#                ball[2])


app = App(title="Results", layout="auto", height=320, width=480, bg="#424242",visible=True)
title = Text(app,str(player)+"'s Results for "+str(drillType),size=20,font="Calibri Bold", color="white")
subt = "Pass Difficulty: "+str(difficulty)+"  |  Ball Speed: "+str(drillSpeed)
subtitle = Text(app,subt,size=14,font="Calibri Bold", color="white")
spacer = Box(app,width=20,height=20)

for i in range(5):

    ball[i+1] = "Pass # "+str(i+1)+":   "+ str((targetTimes[i+1])/1000)+" sec  |  "+str(ballSpeeds[i+1])+" m/s"
    # if len(ball[i+1])!=25:
        # ball[i+1] += "0"
    Text(app,ball[i+1],size=12,font="Calibri Bold", color="white")

app.display()


