from openpyxl import load_workbook, Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from tkinter import *
from threading import Thread

class OutPut():
    def __init__(self, target_times, ball_speeds, player):
        self.target_times = target_times
        self. ball_speeds = ball_speeds
        self.player = player
        self.ball = [0,0,0,0,0,0]
        self.filename = "/home/treen/Documents/School/5.Winter 2019/4A06 S.T.A.R.S/STARS_PROGRAM/STARS/include/demo_wb.xlsx"

    def create_excel_book(self):
        try:
            wb = load_workbook(filename=self.filename)
            print("[Launcher]: file exists")
            new_book = False
        except Exception as e:
            print("[Launcher]: file not found, making new book ")
            wb = Workbook()
            summary_sheet = wb.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Summary'
            new_book = True

        try:
            players_sheet = wb[self.player]
            print("[Launcher]: sheet exists, appending data")

        except KeyError as k:
            print("[Launcher]: creating new sheet for: ", self.player)
            players_sheet = wb.create_sheet(title=self.player)

        data_printed = False
        col = 1
        row = 1
        drill = 1
        while not data_printed:
            if players_sheet.cell(column=col, row=row + 1).value is None:
                players_sheet.cell(column=col, row=row + 1).value = "Drill" + str(drill)
                if row == 1:
                    # players_sheet.cell(column=col, row=row).value = "."
                    players_sheet.cell(column=col + 1, row=row).value = "Ball #"
                    players_sheet.cell(column=col + 2, row=row).value = "Time for Pass"
                    players_sheet.cell(column=col + 3, row=row).value = "Ball Speed"
                for i in range(5):
                    players_sheet.cell(column=col + 1, row=row + i + 1).value = row + i
                    players_sheet.cell(column=col + 2, row=row + i + 1).value = (self.target_times[i + 1] / 1000)
                    players_sheet.cell(column=col + 3, row=row + i + 1).value = self.ball_speeds[i + 1]

                data_printed = True

            else:
                row += 5
                drill += 1

        wb.save(self.filename)
        # PLOT THE DATA:
        # player_sheets =
        time_ranges = [0] * (len(wb.worksheets) - 1)
        # speed_ranges = [0] * (len(wb.worksheets) - 1)

        time_series = [0] * (len(wb.worksheets) - 1)
        # speed_series = [0] * (len(wb.worksheets) - 1)


        for i in range((len(wb.worksheets) - 1)):
            print(str(wb.worksheets[i+1])[12:19])
            gathered = False
            coll = 1
            roww = 1
            chart = ScatterChart()
            timecharts = [chart] * (len(wb.worksheets) - 1)
            timecharts[i].title = str(wb.worksheets[i+1])[12:19] + "'s  Results"
            # speedcharts = [chart]* (len(wb.worksheets) - 1)
            # speedcharts[i].title = str(wb.worksheets[i+1])[12:19] + "'s  Results"
            while not gathered:
                print("Gathering..")
                if wb.worksheets[i + 1].cell(column=coll, row=roww + 1).value is None:
                    gathered = True
                else:
                    roww += 5
                    gathered = False

            time_ranges[i] = Reference(wb.worksheets[i + 1], min_col=3, min_row=1, max_col=4, max_row=roww)
            # speed_ranges[i] = Reference(wb.worksheets[i + 1], min_col=4, min_row=1, max_row=roww)

            x_vals = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)

            time_series[i] = Series(time_ranges[i], xvalues=x_vals, title_from_data=True)
            # speed_series[i] = Series(speed_ranges[i], xvalues=x_vals, title_from_data=True)
            # chart.add_data(time_ranges[i],titles_from_data=True)
            timecharts[i].append(time_series[i])
            timecharts[i].x_axis.title = "Test Number"
            timecharts[i].x_axis.scaling.min = 0
            timecharts[i].x_axis.scaling.max = roww
            timecharts[i].y_axis.title = "Reaction Times (seconds)"
            timecharts[i].legend = None

            s1 = timecharts[i].time_series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline

            s1.graphicalProperties.line.noFill = True

            # speedcharts[i].append(speed_series[i])
            # speedcharts[i].x_axis.title = "Test Number"
            # speedcharts[i].x_axis.scaling.min = 0
            # speedcharts[i].x_axis.scaling.max = roww
            # speedcharts[i].y_axis.title = "Passing Speed (m/s)"
            # speedcharts[i].legend = None
            #
            # s2 = speedcharts[i].speed_series[0]
            # s2.marker.symbol = "triangle"
            # s2.marker.graphicalProperties.solidFill = "FFFF00"  # Marker filling
            # s2.marker.graphicalProperties.line.solidFill = "FFFF00"  # Marker outline
            #
            # s2.graphicalProperties.line.noFill = True
            # s2 = chart.time_series[1]
            # s2.marker.symbol = "circle"
            # s2.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
            # s2.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
            #
            # s2.graphicalProperties.line.noFill = True


            timecharts[i].style = 13
            wb.worksheets[0].add_chart(timecharts[i], "A" + str(i + 1 + 15 * i))
            # wb.worksheets[0].add_chart(speedcharts[i], "J" + str(i + 1 + 15 * i))

        wb.save(self.filename)



    def display_data(self):
        for i in range(5):
            self.ball[i + 1] = "Pass # " + str(i + 1) + ":   " + str((self.target_times[i + 1]) / 1000) + " sec  |  " + str(
                self.ball_speeds[i + 1]) + " m/s"

        results = Tk()
        var1 = StringVar()
        var2 = StringVar()
        var3 = StringVar()
        var4 = StringVar()
        var5 = StringVar()

        ball_1 = Message(results, width=500, textvariable=var1, padx=30,pady=10, anchor=W)
        ball_2 = Message(results, width=500, textvariable=var2, padx=30,pady=10, anchor=W)
        ball_3 = Message(results, width=500, textvariable=var3, padx=30,pady=10, anchor=W)
        ball_4 = Message(results, width=500, textvariable=var4, padx=30,pady=10, anchor=W)
        ball_5 = Message(results, width=500, textvariable=var5, padx=30,pady=10, anchor=W)

        var1.set(self.ball[1])
        var2.set(self.ball[2])
        var3.set(self.ball[3])
        var4.set(self.ball[4])
        var5.set(self.ball[5])

        ball_1.pack()
        ball_2.pack()
        ball_3.pack()
        ball_4.pack()
        ball_5.pack()

        results.mainloop()

    def run(self):
        self.create_excel_book()
        display_thread = Thread(target=self.display_data()).start()


if __name__ == "__main__":
    from multiprocessing import Queue
    import outputfile

    targetTimes = [3000, 4500, 2960, 6200, 5754, 7430]
    ballSpeeds = [4.44, 2.52, 4.32, 3.2, 5.4, 10.2]
    filename = "demo_wb.xlsx"
    player = "Player1"
    difficulty = 1
    drillSpeed = 1
    drillType = "Dynamic"
    ball = [0, 0, 0, 0, 0, 0]
    data = Queue()

    # result = targetTimes + ballSpeeds
    # print(result)
    # data.put(result)
    #
    # out = data.get()
    # print(out)
    OutPut(targetTimes, ballSpeeds, player).run()
